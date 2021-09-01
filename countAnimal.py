from openpyxl import Workbook, load_workbook
wb = load_workbook('animal.xlsx')

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
# ws['A1'] = 42
ws.title = "IamWorksheet"

# Rows can also be appended
# ws.append([1, 2, 3])

for x in range( 1, 4):
    for y in range (2, 5):
        print(ws.cell(x,y).value)

ws["E1"].value = "=sum(B1:D1)"

# Save the file
wb.save("animal.xlsx")