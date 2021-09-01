from openpyxl import Workbook, load_workbook
wb = load_workbook('sample.xlsx')

# grab the active worksheet
ws = wb.active

print(ws["A1"].value)

ws["A1"].value = 55

# Data can be assigned directly to cells
# ws['A1'] = 42
ws.title = "IamWorksheet"

# Rows can also be appended
# ws.append([1, 2, 3])

for x in range( 1, 100):
    for y in range (1, 100):
        ws.cell(x,y, x*y)

# Save the file
wb.save("sample.xlsx")